from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
from openpyxl import load_workbook
import os
import glob

driver = webdriver.Edge()

wb = load_workbook("ser_mente.xlsx")
sheet = wb.active

driver.get('https://sermenteeditorial.com.br/wp-login.php?redirect_to=https%3A%2F%2Fsermenteeditorial.com.br%2Fwp-admin%2F&reauth=1')

driver.maximize_window()

WebDriverWait(driver,60).until(
    EC.presence_of_element_located((By.XPATH,'/html/body/div[1]/form/p[1]/input'))
).send_keys('tatimfmesquita@gmail.com')

driver.find_element(By.XPATH,'/html/body/div[1]/form/div/div/input').send_keys('tati140675')
driver.find_element(By.XPATH,'/html/body/div[1]/form/p[3]/input[1]').click()

link = WebDriverWait(driver,60).until(
    EC.presence_of_element_located((By.XPATH,'/html/body/div[1]/div[1]/div[2]/ul/li[13]/ul/li[3]/a'))
)
link = link.get_attribute('href')
driver.get(link)

elementos = WebDriverWait(driver,60).until(
    EC.presence_of_all_elements_located((By.XPATH,"//tr[td[@data-colname='Page']//a[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'antologia')]]/td[1]/a"))
)
lista_link = []
for elemento in elementos:
    link_view = elemento.get_attribute('href')
    lista_link.append(link_view)

for link in lista_link:
    driver.switch_to.new_window('tab') 
    driver.get(link)
    linhaf = sheet.max_row
    
    sheet.cell(row=linhaf+1,column=1,value=WebDriverWait(driver,60).until(EC.presence_of_element_located((By.XPATH,"//td[@class='' and text()='Nome/Pseudônimo']/following-sibling::td[1]"))).text)
    sheet.cell(row=linhaf+1,column=2,value= driver.find_element(By.XPATH,"//td[@class='' and text()='Email']/following-sibling::td[1]").text)
    sheet.cell(row=linhaf+1,column=3,value=driver.find_element(By.XPATH,"//td[@class='' and text()='Endereço (logradouro']/following-sibling::td[1]").text)
    sheet.cell(row=linhaf+1,column=4,value=driver.find_element(By.XPATH,"//td[@class='' and text()='Bairro']/following-sibling::td[1]").text)
    sheet.cell(row=linhaf+1,column=5,value=driver.find_element(By.XPATH,"//td[@class='' and text()='Cidade']/following-sibling::td[1]").text)
    sheet.cell(row=linhaf+1,column=6,value=driver.find_element(By.XPATH,"//td[@class='' and text()='Estado']/following-sibling::td[1]").text)
    sheet.cell(row=linhaf+1,column=7,value=driver.find_element(By.XPATH,"//td[@class='' and text()='País']/following-sibling::td[1]").text)
    sheet.cell(row=linhaf+1,column=8,value=driver.find_element(By.XPATH,"//td[@class='' and text()='CPF']/following-sibling::td[1]").text)
    sheet.cell(row=linhaf+1,column=9,value=driver.find_element(By.XPATH,"//td[@class='' and text()='Celular de contato']/following-sibling::td[1]").text)
    sheet.cell(row=linhaf+1,column=10,value=driver.find_element(By.XPATH,"//td[@class='' and text()='Título do texto a ser apresentado']/following-sibling::td[1]").text)
    sheet.cell(row=linhaf+1,column=11,value=(driver.find_element(By.XPATH,"/html/body/div/div[2]/div[3]/div[1]/div[2]/div/div/div/div/form/div[1]/div[1]/div/div[1]/div/h2").text.split('Submission #')[1]))
    pdf = driver.find_element(By.XPATH,'/html/body/div/div[2]/div[3]/div[1]/div[2]/div/div/div/div/form/div[1]/div[1]/div/div[2]/table/tbody/tr[14]/td[2]/div/a')
    pdf = pdf.get_attribute('href')
    driver.get(pdf)
    diretorio_downloads = os.path.expanduser('~' + os.sep + 'Downloads')

    sleep(1)
    
    diretorio_downloads = os.path.expanduser('~' + os.sep + 'Downloads')


    todos_arquivos = glob.glob(os.path.join(diretorio_downloads, '*'))




   
    arquivo_mais_recente = max(todos_arquivos, key=os.path.getctime)

    
    novo_diretorio = 'C:\Comprovantes'

   
    os.makedirs(novo_diretorio, exist_ok=True)

    
    novo_caminho = os.path.join(novo_diretorio, os.path.basename(arquivo_mais_recente))

    
    os.rename(arquivo_mais_recente, novo_caminho)
    
    
    

wb.save("ser_mente_final.xlsx")
driver.quit()
